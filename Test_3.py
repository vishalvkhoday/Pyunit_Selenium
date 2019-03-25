'''
Created on Jan 4, 2019

@author: vkhoday
'''

from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import HTMLTestRunner
from HtmlTestRunner.result import TestResult
from _datetime import datetime
import unittest
import sys
import time


# from webbrowser import Chrome
# import xmlrunner.xmlrunner

driver = webdriver.Chrome('../Pyunit_Selenium/WebDriver/chromedriver_235')
# driver_RC =webdriver.Remote(desired_capabilities={"browserName" :Chrome,"platform":"WINDOWS","node":5555})


def login():
    driver.get('https://www.nseindia.com')
    driver.maximize_window()
    driver.implicitly_wait(5)
#     webTitle = WebDriverWait(driver,5).until(EC.element_to_be_clickable(By.XPATH,'//*[@id="keyword"]'))
    
    
def ObjWorkBook(File_name):
    ObjWb = load_workbook(File_name)
    return ObjWb


def GetRowCountXl(strFile_name,strSheetName):
    Wb=ObjWorkBook(strFile_name)
    Ws=Wb[strSheetName]
    rCnt = Ws.max_row
    Wb.close()
    return rCnt

def broswer_close():
    driver.quit()
    
def GetScriptName(File_name,SheetName):
    Wb= ObjWorkBook(File_name)
    Ws =Wb[SheetName] 
    scrlist ={}
    rCnt = Ws.max_row
    for r in range(2,rCnt):
        scr = Ws['A' + str(r)].value
        exe_St = Ws['N' + str(r)].value
        try:
            scrlist.update({scr:exe_St})
        except:
            Wb.close()
        
    Wb.close()
    return scrlist
    
def Script_input():
    driver.find_element_by_xpath('//*[@class="search"]/input[@id="keyword"]').clear()
    driver.find_element_by_xpath('//*[@class="search"]/input[@id="keyword"]').send_keys("Shantigear")
    driver.set_page_load_timeout(4)
    time.sleep(3)

def Result_gen(self):
    try:
#         HTMLTestRunner.TestResult.addSuccess(self,scr_nm)
#         HTMLTestRunner.TestResult.wasSuccessful()
        print("Test result function")
        
        
#                     unittest.TestResult.addSuccess(self, scr_nm)
#                     HTMLTestRunner.HTMLTestRunner.title('TestReports')
#                     self.stream.writeln("OK")
#                     HtmlTestRunner.result.load_template(template)
    except Exception as e:
        print (e.sys.exc_info()[0])
        HTMLTestRunner.TestResult.addFailure('Error')
    



