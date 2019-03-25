from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options 
import time
import pyautogui
import openpyxl 
from test import test_errno
import os



driver = webdriver.Chrome(executable_path='C:/Python36/chromedriver_235')
f_Path = 'C:/Users/vkhoday/git/Selenium_NSE_Algo/Additonal_Utility/NSE_Script_codes26Dec2018.xlsx'


url = "https://www.moneycontrol.com/"
home_scr_txt = '//*[@id="search_str"]'
#home_link = '//*[@id="newsn"]/div/div[2]/p/a'




def get_AbsPath():
    path_scr = os.getcwd()
    return (path_scr)

# driver = webdriver.Chrome(executable_path =temp_path+'/WebDriver/chromedriver_235.exe')
temp_path = get_AbsPath()

def get_Script_name():
    f_Path = 'C:/Users/vkhoday/git/Selenium_NSE_Algo/Additonal_Utility/Down_Scripts_List.xlsx'
    Wb = openpyxl.load_workbook(f_Path)
    Ws =Wb['Sheet1']
    r_count = Ws.max_row
    dic_script = {}
    for i in range(2,r_count):
        colA = 'A' + str(i)
        colB = 'N'+ str(i)
        script_name = Ws[colA].value
        status = Ws[colB].value
        dic_script.update({script_name:status})
        
    print (dic_script)
    Wb.close()
    
    return dic_script
        

def Screenshot():
        timstr = str(time.strftime('%Y%m%d%H%M%S'))
        pic =pyautogui.screenshot() #,)
        pic.save('Screen_shot/img'+timstr+'.png')
        
def get_screenShot():
    temp_path=get_AbsPath()
    driver.get_screenshot_as_file(temp_path+'/Screen_shot/'+str(time.strftime('%Y%m%d%H%M%S'))+'.png')

# def Screenshot():
#         timstr = str(time.strftime('%Y%m%d%H%M%S'))
#         driver.get_screenshot_as_file(path_scr+'/Screen_shot/'+timstr+'.png')
# #         pic.save('Screen_shot\\img'+timstr+'.png')

        
        
def temp_scr_shot():
    driver.get_screenshot_as_png()

def closeExcel():
    Wb = openpyxl.load_workbook(f_Path)
    Wb.close()
    
    
    

